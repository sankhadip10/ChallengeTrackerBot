import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def generate_excel(eligible_users):
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eligible Participants"

    # Styling
    header_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    fill = PatternFill(start_color="E5E4E2", end_color="E5E4E2", fill_type="solid")

    # Headers
    headers = ['User ID', 'Streak', 'Last Post Date']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].border = border
        ws[f'{col_letter}1'].fill = fill
        ws[f'{col_letter}1'].alignment = Alignment(horizontal="center")

        # Setting column width (just an arbitrary width for all columns here, adjust as needed)
        ws.column_dimensions[col_letter].width = 20

    # Populate data
    for row_num, user_data in enumerate(eligible_users, 2):  # Start from second row as headers are in the first
        ws.cell(row=row_num, column=1, value=user_data["user_id"]).border = border
        ws.cell(row=row_num, column=2, value=user_data["streak"]).border = border
        ws.cell(row=row_num, column=3, value=user_data["last_post_date"]).border = border

    # Save to file
    filename = "eligible_participants.xlsx"
    wb.save(filename)
    print(f"{filename} has been saved!")
